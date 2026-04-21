"""
Admin panel handlers for Mars Intern Bot
"""
from datetime import date, timedelta, datetime, time
from aiogram import F, Router, Bot
from aiogram.types import Message, InlineKeyboardMarkup, InlineKeyboardButton, CallbackQuery, FSInputFile
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.filters import CommandStart

from database import db
from interns import INTERNS
from keyboards import get_main_keyboard


class AdminStates(StatesGroup):
    """Admin states for FSM"""
    searching_intern = State()
    deleting_report = State()
    adding_admin = State()
    writing_issue_reason = State()  # For reporting issues

admin_router = Router()

# Admin user IDs (will be stored in database)
# Can be set via /admin_add command


def get_back_keyboard() -> InlineKeyboardMarkup:
    """Get keyboard with back button"""
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🔙 Orqaga", callback_data="admin_back")]
    ])


@admin_router.message(F.text == "/admin")
async def admin_menu(message: Message):
    """Show admin menu with inline buttons"""
    if not db.is_admin(message.from_user.id):
        await message.answer("❌ Siz admin emassiz!")
        return
    
    # Create inline keyboard with buttons
    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="📊 Statistika", callback_data="admin_stats"),
            InlineKeyboardButton(text="📋 Hisobotlar", callback_data="admin_reports")
        ],
        [
            InlineKeyboardButton(text="📚 Dars ro'yxati", callback_data="admin_lessons"),
            InlineKeyboardButton(text="👥 Talabalar", callback_data="admin_interns")
        ],
        [
            InlineKeyboardButton(text="⏱️ Ish vaqti", callback_data="admin_work_stats"),
            InlineKeyboardButton(text="📥 Excel export", callback_data="admin_excel")
        ],
        [
            InlineKeyboardButton(text="🔍 Qidirish", callback_data="admin_search")
        ],
        [
            InlineKeyboardButton(text="🗑️ O'chirish", callback_data="admin_delete"),
            InlineKeyboardButton(text="👨‍💼 Admin qo'shish", callback_data="admin_add")
        ],
        [
            InlineKeyboardButton(text="ℹ️ Yordam", callback_data="admin_help")
        ]
    ])
    
    menu_text = """
🔐 ADMIN PANEL

Siz quyidagi amalarni bajarishingiz mumkin:
1️⃣ 📊 Statistika - Bugungi statistikani ko'rish
2️⃣ 📋 Hisobotlar - Bugungi hisobotlarni ko'rish
3️⃣ 📚 Dars ro'yxati - Interns kiritgan darslarni ko'rish
4️⃣ 👥 Talabalar - Barcha talabalari ko'rish
5️⃣ ⏱️ Ish vaqti - Har bir intern nechi soat marsda bo'gani
6️⃣ 📥 Excel export - Darslarni Excel ga chiqarish
7️⃣ 📜 Jurnali - Faoliyat jurnalini ko'rish
8️⃣ 🔍 Qidirish - Talabani qidirish
9️⃣ 🗑️ O'chirish - Hisobotni o'chirish
"""
    
    await message.answer(menu_text, reply_markup=keyboard)


@admin_router.message(F.text == "/stats")
async def admin_stats(message: Message):
    """Show today's statistics"""
    if not db.is_admin(message.from_user.id):
        await message.answer("❌ Siz admin emassiz!")
        return
    
    today = date.today()
    summary = db.get_attendance_summary(today)
    
    stats_text = f"""
📊 BUGUNGI STATISTIKA
📅 Sana: {today.strftime('%d.%m.%Y')}

👥 Jami talabalar: {len(INTERNS)}
✅ Kelganlar: {summary['present']}
❌ Kelmaganlar: {summary['absent']}
⏳ Hisobota kutayotganlar: {summary['not_reported']}

Kelish foizi: {(summary['present'] / len(INTERNS) * 100):.1f}%
"""
    
    await message.answer(stats_text)
    db.add_log(message.from_user.id, "stats_viewed", f"Date: {today}")


@admin_router.callback_query(F.data == "admin_stats")
async def admin_stats_callback(query: CallbackQuery):
    """Handle statistics button click"""
    if not db.is_admin(query.from_user.id):
        await query.answer("❌ Siz admin emassiz!", show_alert=True)
        return
    
    today = date.today()
    summary = db.get_attendance_summary(today)
    
    stats_text = f"""
📊 BUGUNGI STATISTIKA
📅 Sana: {today.strftime('%d.%m.%Y')}

👥 Jami talabalar: {len(INTERNS)}
✅ Kelganlar: {summary['present']}
❌ Kelmaganlar: {summary['absent']}
⏳ Hisobota kutayotganlar: {summary['not_reported']}

Kelish foizi: {(summary['present'] / len(INTERNS) * 100):.1f}%
"""
    
    await query.message.edit_text(stats_text, reply_markup=get_back_keyboard())
    await query.answer()
    db.add_log(query.from_user.id, "stats_viewed", f"Date: {today}")


@admin_router.callback_query(F.data == "admin_reports")
async def admin_reports_callback(query: CallbackQuery):
    """Show today's reports"""
    if not db.is_admin(query.from_user.id):
        await query.answer("❌ Siz admin emassiz!", show_alert=True)
        return
    
    today = date.today()
    reports = db.get_reports_by_date(today)
    
    if not reports:
        await query.answer(f"📭 {today.strftime('%d.%m.%Y')} uchun hisobotlar yo'q")
        return
    
    reports_text = f"📋 BUGUNGI HISOBOTLAR ({today.strftime('%d.%m.%Y')})\n\n"
    
    for report in reports:
        status = "✅ Keldi" if report['status'] == 'Keldi' else "❌ Kelmadi"
        reports_text += f"{status} {report['intern_name']}\n"
        
        if report['status'] == 'Keldi':
            reports_text += f"  🕒 {report['arrival_time']} - {report['departure_time']}\n"
            reports_text += f"  📚 {report['lesson_count']} dars: {report['teachers']}\n"
        else:
            reports_text += f"  Sabab: {report['absence_reason']}\n"
        reports_text += "\n"
    
    await query.message.edit_text(reports_text, reply_markup=get_back_keyboard())
    await query.answer()
    db.add_log(query.from_user.id, "reports_viewed", f"Date: {today}")


@admin_router.callback_query(F.data == "admin_lessons")
async def admin_lessons_callback(query: CallbackQuery):
    """Show lessons list for last 30 days"""
    if not db.is_admin(query.from_user.id):
        await query.answer("❌ Siz admin emassiz!", show_alert=True)
        return
    
    # Get today's lessons first
    today_lessons = db.get_lessons_by_date(date.today())
    
    if today_lessons:
        # Show today's lessons
        lessons_text = f"📚 BUGUNGI KIRITILGAN DARSLAR\n📅 {date.today().strftime('%d.%m.%Y')}\n\n"
        
        current_intern = None
        for lesson in today_lessons:
            # Group by intern
            if lesson['intern_name'] != current_intern:
                current_intern = lesson['intern_name']
                lessons_text += f"👤 {current_intern}\n"
            
            # Lesson details
            lessons_text += f"   🔹 Dars #{lesson['lesson_number']}: {lesson['teacher_name']}\n"
            if lesson['room']:
                lessons_text += f"      📍 {lesson['room']} | ⏰ {lesson['lesson_time']}\n"
            else:
                lessons_text += f"      ⏰ {lesson['lesson_time']}\n"
        
        await query.message.edit_text(lessons_text, reply_markup=get_back_keyboard())
        await query.answer()
        db.add_log(query.from_user.id, "lessons_viewed", f"Today: {len(today_lessons)} dars")
    else:
        # Try to get last 30 days
        all_lessons = db.get_all_lessons(days=30)
        
        if not all_lessons:
            await query.answer("📭 Darslar topilmadi (oxirgi 30 kun)")
            return
        
        lessons_text = f"📚 KIRITILGAN DARSLAR (Oxirgi 30 kun)\n\n"
        
        current_date = None
        current_intern = None
        
        for lesson in all_lessons:
            # Group by date
            if lesson['lesson_date'] != current_date:
                current_date = lesson['lesson_date']
                lessons_text += f"\n📅 {current_date}\n"
                current_intern = None
            
            # Group by intern
            if lesson['intern_name'] != current_intern:
                current_intern = lesson['intern_name']
                lessons_text += f"   👤 {current_intern}\n"
            
            # Lesson details
            lessons_text += f"      🔹 Dars #{lesson['lesson_number']}: {lesson['teacher_name']}\n"
            if lesson['room']:
                lessons_text += f"         📍 {lesson['room']} | ⏰ {lesson['lesson_time']}\n"
            else:
                lessons_text += f"         ⏰ {lesson['lesson_time']}\n"
        
        # Limit message length
        if len(lessons_text) > 4000:
            lessons_text = lessons_text[:3950] + "\n\n... (Batafsil Excel ga qarang)"
        
        await query.message.edit_text(lessons_text, reply_markup=get_back_keyboard())
        await query.answer()
        db.add_log(query.from_user.id, "lessons_viewed", f"Total: {len(all_lessons)} dars")


@admin_router.callback_query(F.data == "admin_work_stats")
async def admin_work_stats_callback(query: CallbackQuery):
    """Show work sessions statistics"""
    if not db.is_admin(query.from_user.id):
        await query.answer("❌ Siz admin emassiz!", show_alert=True)
        return
    
    today = date.today()
    work_sessions = db.get_work_sessions_by_date(today)
    
    if not work_sessions:
        await query.answer("📭 Bugun ish sessiyalari yo'q")
        return
    
    work_text = f"⏱️ ISH VAQTI STATISTIKASI\n📅 {today.strftime('%d.%m.%Y')}\n\n"
    
    # Group by intern
    intern_data = {}
    intern_totals = {}
    for session in work_sessions:
        intern = session['intern_name']
        if intern not in intern_data:
            intern_data[intern] = []
            intern_totals[intern] = {'completed': 0, 'active': False}
        intern_data[intern].append(session)
        
        # Calculate totals
        if session['status'] == 'completed' and session.get('duration_minutes'):
            intern_totals[intern]['completed'] += session['duration_minutes']
        elif session['status'] == 'active':
            intern_totals[intern]['active'] = True
    
    # Display summary first
    work_text += "📊 QISQA XULOSA\n"
    for intern_name in sorted(intern_totals.keys()):
        total_min = intern_totals[intern_name]['completed']
        is_active = intern_totals[intern_name]['active']
        hours = total_min // 60
        minutes = total_min % 60
        
        if is_active:
            work_text += f"   🟢 {intern_name}: {hours}h {minutes}m + (davom...)\n"
        else:
            work_text += f"   ⚪ {intern_name}: {hours}h {minutes}m\n"
    
    work_text += "\n" + "="*40 + "\n\n"
    work_text += "📝 BATAFSIL\n\n"
    
    # Display grouped data
    for intern_name in sorted(intern_data.keys()):
        sessions = intern_data[intern_name]
        total_minutes = 0
        
        work_text += f"👤 {intern_name}\n"
        
        for session in sessions:
            if session['status'] == 'completed' and session.get('duration_minutes'):
                hours = session['duration_minutes'] // 60
                minutes = session['duration_minutes'] % 60
                
                # Extract time from timestamp
                start = session['start_time']
                end = session.get('end_time', '')
                
                # Handle both datetime strings and regular time strings
                if isinstance(start, str):
                    start_time = start.split('T')[-1][:5] if 'T' in start else start[:5]
                else:
                    start_time = str(start)[:5]
                
                if isinstance(end, str):
                    end_time = end.split('T')[-1][:5] if 'T' in end else end[:5]
                else:
                    end_time = str(end)[:5] if end else ''
                
                work_text += f"   🕐 {start_time} - {end_time} | {hours}h {minutes}m\n"
                total_minutes += session['duration_minutes']
            elif session['status'] == 'active':
                start = session['start_time']
                if isinstance(start, str):
                    start_time = start.split('T')[-1][:5] if 'T' in start else start[:5]
                else:
                    start_time = str(start)[:5]
                work_text += f"   🟢 {start_time} - ... (Davom etayapti)\n"
        
        if total_minutes > 0:
            total_hours = total_minutes // 60
            total_mins = total_minutes % 60
            work_text += f"   📊 Jami: {total_hours}h {total_mins}m ({total_minutes} min)\n"
        work_text += "\n"
    
    # Limit message length
    if len(work_text) > 4000:
        work_text = work_text[:3950] + "\n\n... (Batafsil uchun ro'yxatni qarang)"
    
    await query.message.edit_text(work_text, reply_markup=get_back_keyboard())
    await query.answer()
    db.add_log(query.from_user.id, "work_stats_viewed", f"Date: {today}")


@admin_router.callback_query(F.data == "admin_interns")
async def admin_interns_callback(query: CallbackQuery):
    """Show all interns list"""
    if not db.is_admin(query.from_user.id):
        await query.answer("❌ Siz admin emassiz!", show_alert=True)
        return
    
    interns_text = "👥 BARCHA TALABALAR\n\n"
    for i, intern in enumerate(INTERNS, 1):
        interns_text += f"{i}. {intern}\n"
    
    await query.message.edit_text(interns_text, reply_markup=get_back_keyboard())
    await query.answer()
    db.add_log(query.from_user.id, "interns_list_viewed")


@admin_router.callback_query(F.data == "admin_excel")
async def admin_excel_callback(query: CallbackQuery, bot: Bot):
    """Export data to Excel with beautiful formatting"""
    if not db.is_admin(query.from_user.id):
        await query.answer("❌ Siz admin emassiz!", show_alert=True)
        return
    
    try:
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from database import DATABASE_FILE
        
        today = date.today()
        today_lessons = db.get_lessons_by_date(today)
        lessons_by_intern = {}

        for lesson in today_lessons:
            lessons_by_intern.setdefault(lesson['intern_name'], []).append(lesson.get('teacher_name', '-'))
        
        # ===== SHEET 1: ATTENDANCE REPORT =====
        all_reports = db.get_reports_by_date(today)
        attendance_data = []
        
        for intern in INTERNS:
            report_status = None
            for report in all_reports:
                if report['intern_name'] == intern:
                    report_status = report
                    break
            
            if report_status:
                if report_status['status'] == 'Keldi':
                    teacher_names = lessons_by_intern.get(intern, [])
                    attendance_data.append({
                        'No': len(attendance_data) + 1,
                        'ISM FAMILIYA': intern,
                        'STATUS': 'Keldi',
                        'KELDI': 'Ha',
                        'KELMADI': "Yo'q",
                        'KELISH VAQTI': report_status.get('arrival_time', '-'),
                        'KETISH VAQTI': report_status.get('departure_time', '-'),
                        'DARSLAR': report_status.get('lesson_count', 0),
                        'USTOZ 1': teacher_names[0] if len(teacher_names) > 0 else '-',
                        'USTOZ 2': teacher_names[1] if len(teacher_names) > 1 else '-',
                        'USTOZ 3': teacher_names[2] if len(teacher_names) > 2 else '-',
                        'USTOZ 4': teacher_names[3] if len(teacher_names) > 3 else '-',
                        'USTOZ 5': teacher_names[4] if len(teacher_names) > 4 else '-',
                        'SABAB / IZOH': '-'
                    })
                else:
                    attendance_data.append({
                        'No': len(attendance_data) + 1,
                        'ISM FAMILIYA': intern,
                        'STATUS': 'Kelmadi',
                        'KELDI': "Yo'q",
                        'KELMADI': 'Ha',
                        'KELISH VAQTI': '-',
                        'KETISH VAQTI': '-',
                        'DARSLAR': 0,
                        'USTOZ 1': '-',
                        'USTOZ 2': '-',
                        'USTOZ 3': '-',
                        'USTOZ 4': '-',
                        'USTOZ 5': '-',
                        'SABAB / IZOH': report_status.get('absence_reason', 'Sabab ko\'rsatilmadi')
                    })
            else:
                attendance_data.append({
                    'No': len(attendance_data) + 1,
                    'ISM FAMILIYA': intern,
                    'STATUS': 'Kutilmoqda',
                    'KELDI': '-',
                    'KELMADI': '-',
                    'KELISH VAQTI': '-',
                    'KETISH VAQTI': '-',
                    'DARSLAR': 0,
                    'USTOZ 1': '-',
                    'USTOZ 2': '-',
                    'USTOZ 3': '-',
                    'USTOZ 4': '-',
                    'USTOZ 5': '-',
                    'SABAB / IZOH': 'Hisobot topshirmagan'
                })
        
        # ===== SHEET 2: LESSONS =====
        lessons = db.get_all_lessons(days=90)
        lessons_data = []
        
        for lesson in lessons:
            lessons_data.append({
                'Sana': lesson['lesson_date'],
                'ISM': lesson['intern_name'],
                'Dars #': lesson['lesson_number'],
                'Ustoz': lesson['teacher_name'],
                'Xona': lesson.get('room', '-'),
                'Vaqt': lesson['lesson_time']
            })
        
        # Create excel file
        excel_path = DATABASE_FILE.parent / f"hisobot_{today.strftime('%d_%m_%Y')}.xlsx"
        
        # Create writer
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            # Write attendance sheet
            df_attendance = pd.DataFrame(attendance_data)
            df_attendance.to_excel(writer, sheet_name='Hisobot', index=False)
            
            # Write lessons sheet
            if lessons_data:
                df_lessons = pd.DataFrame(lessons_data)
                df_lessons.to_excel(writer, sheet_name='Darslar', index=False)
        
        # Format Excel file with beautiful styling
        wb = load_workbook(excel_path)

        # Format Hisobot sheet
        ws_hisobot = wb['Hisobot']

        present_count = sum(1 for row in attendance_data if row['STATUS'] == 'Keldi')
        absent_count = sum(1 for row in attendance_data if row['STATUS'] == 'Kelmadi')
        pending_count = sum(1 for row in attendance_data if row['STATUS'] == 'Kutilmoqda')

        # Add title and summary block
        ws_hisobot.insert_rows(1, amount=3)
        ws_hisobot['A1'] = f"MARS INTERN DAVOMAT HISOBOTI - {today.strftime('%d.%m.%Y')}"
        ws_hisobot.merge_cells("A1:M1")

        ws_hisobot['A2'] = "Jami internlar"
        ws_hisobot['B2'] = len(INTERNS)
        ws_hisobot['C2'] = "Keldi"
        ws_hisobot['D2'] = present_count
        ws_hisobot['E2'] = "Kelmadi"
        ws_hisobot['F2'] = absent_count
        ws_hisobot['G2'] = "Kutilmoqda"
        ws_hisobot['H2'] = pending_count
        ws_hisobot['I2'] = f"Kelish foizi: {(present_count / len(INTERNS) * 100):.1f}%"

        title_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        title_font = Font(bold=True, color='FFFFFF', size=14)
        summary_label_fill = PatternFill(start_color='D9EAF7', end_color='D9EAF7', fill_type='solid')
        summary_value_fill = PatternFill(start_color='F4F8FB', end_color='F4F8FB', fill_type='solid')
        ws_hisobot['A1'].fill = title_fill
        ws_hisobot['A1'].font = title_font
        ws_hisobot['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws_hisobot.row_dimensions[1].height = 28
        ws_hisobot.row_dimensions[2].height = 22

        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col_num in range(1, 10):
            cell = ws_hisobot.cell(row=2, column=col_num)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if col_num % 2 == 1:
                cell.fill = summary_label_fill
                cell.font = Font(bold=True, color='1F1F1F')
            else:
                cell.fill = summary_value_fill
                cell.font = Font(bold=True, color='1F1F1F')

        # Header formatting
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        for col_num, column_title in enumerate(df_attendance.columns, 1):
            cell = ws_hisobot.cell(row=4, column=col_num)
            cell.value = column_title
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border

        # Data formatting
        for row_num, row_data in enumerate(attendance_data, 5):
            status_value = row_data['STATUS']
            if status_value == 'Keldi':
                row_fill = PatternFill(start_color='EAF6EA', end_color='EAF6EA', fill_type='solid')
            elif status_value == 'Kelmadi':
                row_fill = PatternFill(start_color='FDE9E7', end_color='FDE9E7', fill_type='solid')
            else:
                row_fill = PatternFill(start_color='FFF4D6', end_color='FFF4D6', fill_type='solid')

            for col_num, col_value in enumerate(row_data.values(), 1):
                cell = ws_hisobot.cell(row=row_num, column=col_num)
                cell.value = col_value
                cell.border = border
                cell.fill = row_fill
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

                if col_num in (1, 3, 4, 5, 6, 7):
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                if col_num == 3:
                    if col_value == 'Keldi':
                        cell.font = Font(color='006100', bold=True)
                    elif col_value == 'Kelmadi':
                        cell.font = Font(color='9C0006', bold=True)
                    else:
                        cell.font = Font(color='9C6500', bold=True)
                elif col_num in (4, 5):
                    cell.font = Font(bold=True)

            ws_hisobot.row_dimensions[row_num].height = 24

        # Set column widths
        ws_hisobot.column_dimensions['A'].width = 5
        ws_hisobot.column_dimensions['B'].width = 24
        ws_hisobot.column_dimensions['C'].width = 14
        ws_hisobot.column_dimensions['D'].width = 10
        ws_hisobot.column_dimensions['E'].width = 10
        ws_hisobot.column_dimensions['F'].width = 14
        ws_hisobot.column_dimensions['G'].width = 14
        ws_hisobot.column_dimensions['H'].width = 24
        ws_hisobot.column_dimensions['I'].width = 24
        ws_hisobot.column_dimensions['J'].width = 24
        ws_hisobot.column_dimensions['K'].width = 24
        ws_hisobot.column_dimensions['L'].width = 24
        ws_hisobot.column_dimensions['M'].width = 34
        ws_hisobot.freeze_panes = 'A5'
        ws_hisobot.auto_filter.ref = f"A4:M{ws_hisobot.max_row}"

        # Format Darslar sheet if it exists
        if 'Darslar' in wb.sheetnames:
            ws_lessons = wb['Darslar']
            
            # Add title
            ws_lessons.insert_rows(1)
            ws_lessons['A1'] = 'DARS RO\'YXATI (Oxirgi 90 kun)'
            ws_lessons.merge_cells('A1:F1')
            
            ws_lessons['A1'].fill = title_fill
            ws_lessons['A1'].font = title_font
            ws_lessons['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws_lessons.row_dimensions[1].height = 25
            
            # Header formatting
            for col_num, column_title in enumerate(df_lessons.columns, 1):
                cell = ws_lessons.cell(row=2, column=col_num)
                cell.value = column_title
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Data formatting
            for row_num, row_data in enumerate(lessons_data, 3):
                for col_num, col_value in enumerate(row_data.values(), 1):
                    cell = ws_lessons.cell(row=row_num, column=col_num)
                    cell.value = col_value
                    cell.border = border
                    cell.alignment = Alignment(horizontal='left', vertical='center')

            ws_lessons.freeze_panes = 'A3'
            ws_lessons.auto_filter.ref = f"A2:F{ws_lessons.max_row}"
            
            # Set column widths
            ws_lessons.column_dimensions['A'].width = 15
            ws_lessons.column_dimensions['B'].width = 20
            ws_lessons.column_dimensions['C'].width = 10
            ws_lessons.column_dimensions['D'].width = 20
            ws_lessons.column_dimensions['E'].width = 15
            ws_lessons.column_dimensions['F'].width = 15
        
        wb.save(excel_path)
        
        # Send file
        file = FSInputFile(excel_path)
        stats_text = f"""
📊 HISOBOT XULOSA
📅 {today.strftime('%d.%m.%Y')}

✅ Kelganlar: {present_count} ta
❌ Kelmaganlar: {absent_count} ta
⏳ Kutilmoqda: {pending_count} ta

📚 Jami darslar: {len(lessons)} ta
"""
        
        await bot.send_document(
            chat_id=query.from_user.id,
            document=file,
            caption=stats_text
        )
        
        await query.message.edit_text(f"✅ Hisobot yaratildi va yuborsdim!\n📄 {excel_path.name}", reply_markup=get_back_keyboard())
        await query.answer()
        db.add_log(query.from_user.id, "excel_exported", f"Attendance + {len(lessons)} lessons")
    except Exception as e:
        error_msg = f"❌ Xato: {str(e)}"
        await query.message.edit_text(error_msg, reply_markup=get_back_keyboard())
        await query.answer()
        print(f"Excel export error: {str(e)}")


@admin_router.callback_query(F.data == "admin_logs")
async def admin_logs_callback(query: CallbackQuery):
    """Show activity logs"""
    if not db.is_admin(query.from_user.id):
        await query.answer("❌ Siz admin emassiz!", show_alert=True)
        return
    
    logs = db.get_logs(limit=20)
    
    logs_text = "📜 FAOLIYAT JURNALI (Oxirgi 20)\n\n"
    
    for log in logs:
        logs_text += f"👤 {log['user_id']}\n"
        logs_text += f"   ▪ {log['action']}\n"
        if log['details']:
            logs_text += f"   ▪ {log['details']}\n"
        logs_text += f"   ⏰ {log['created_at']}\n\n"
    
    await query.message.edit_text(logs_text, reply_markup=get_back_keyboard())
    await query.answer()


@admin_router.callback_query(F.data == "admin_search")
async def admin_search_callback(query: CallbackQuery, state: FSMContext):
    """Search for specific report"""
    if not db.is_admin(query.from_user.id):
        await query.answer("❌ Siz admin emassiz!", show_alert=True)
        return
    
    await query.answer()
    await query.message.answer("🔍 Talabaning ismi ni yozing:")
    await state.set_state(AdminStates.searching_intern)


@admin_router.message(AdminStates.searching_intern)
async def process_search(message: Message, state: FSMContext):
    """Process intern search"""
    intern_name = message.text.strip()
    
    # Check for cancel command
    if intern_name.lower() == "bekor":
        await message.answer("Bekor qilindi.", reply_markup=get_main_keyboard())
        await state.clear()
        return
    
    reports = db.get_reports_by_intern(intern_name, days=30)
    
    if not reports:
        await message.answer(f"❌ {intern_name} uchun hisobotlar topilmadi")
    else:
        search_text = f"🔍 {intern_name} - Oxirgi 30 kun hisobotlari:\n\n"
        
        for report in reports:
            status = "✅" if report['status'] == 'Keldi' else "❌"
            search_text += f"{status} {report['report_date']}\n"
            
            if report['status'] == 'Keldi':
                search_text += f"   {report['arrival_time']} - {report['departure_time']} | {report['lesson_count']} dars\n"
            else:
                search_text += f"   {report['absence_reason']}\n"
            search_text += "\n"
        
        keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🔙 Orqaga", callback_data="admin_back")]
        ])
        
        await message.answer(search_text, reply_markup=keyboard)
        db.add_log(message.from_user.id, "search_report", f"Intern: {intern_name}")
    
    await state.clear()


@admin_router.callback_query(F.data == "admin_delete")
async def admin_delete_callback(query: CallbackQuery, state: FSMContext):
    """Delete report"""
    if not db.is_admin(query.from_user.id):
        await query.answer("❌ Siz admin emassiz!", show_alert=True)
        return
    
    await query.answer()
    await query.message.answer("🗑️ Talabaning ismi va sanani yozing:\nMisol: Humoyun 17.04.2026")
    await state.set_state(AdminStates.deleting_report)


@admin_router.message(AdminStates.deleting_report)
async def process_delete(message: Message, state: FSMContext):
    """Process report deletion"""
    # Check for cancel command
    if message.text.strip().lower() == "bekor":
        await message.answer("Bekor qilindi.", reply_markup=get_main_keyboard())
        await state.clear()
        return
    
    parts = message.text.split(" ", 1)
    if len(parts) < 2:
        await message.answer("❌ Tog'ri format: Ism Familiya DD.MM.YYYY")
        return
    
    intern_name = parts[0]
    date_str = parts[1]
    
    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🔙 Orqaga", callback_data="admin_back")]
    ])
    
    try:
        report_date = date.strptime(date_str, "%d.%m.%Y")
        if db.delete_report(intern_name, report_date):
            result_msg = f"✅ Hisobot o'chirildi: {intern_name} ({date_str})"
            await message.answer(result_msg, reply_markup=keyboard)
            db.add_log(message.from_user.id, "report_deleted", f"{intern_name} on {date_str}")
        else:
            result_msg = "❌ Hisobot o'chirilmadi"
            await message.answer(result_msg, reply_markup=keyboard)
    except ValueError:
        result_msg = "❌ Sana formati xato (DD.MM.YYYY ishlatilsin)"
        await message.answer(result_msg, reply_markup=keyboard)
    
    await state.clear()


@admin_router.callback_query(F.data == "admin_add")
async def admin_add_callback(query: CallbackQuery, state: FSMContext):
    """Add new admin user"""
    if not db.is_admin(query.from_user.id):
        await query.answer("❌ Siz admin emassiz!", show_alert=True)
        return
    
    await query.answer()
    await query.message.answer("👨‍💼 Yangi admin User ID sini yozing:")
    await state.set_state(AdminStates.adding_admin)


@admin_router.message(AdminStates.adding_admin)
async def process_add_admin(message: Message, state: FSMContext):
    """Process adding new admin"""
    # Check for cancel command
    if message.text.strip().lower() == "bekor":
        await message.answer("Bekor qilindi.", reply_markup=get_main_keyboard())
        await state.clear()
        return
    
    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🔙 Orqaga", callback_data="admin_back")]
    ])
    
    try:
        new_admin_id = int(message.text)
        if db.add_admin(new_admin_id):
            result_msg = f"✅ Admin qo'shildi: {new_admin_id}"
            await message.answer(result_msg, reply_markup=keyboard)
            db.add_log(message.from_user.id, "admin_added", f"Admin ID: {new_admin_id}")
        else:
            result_msg = "❌ Admin qo'shilmadi (Balki allaqachon admin)"
            await message.answer(result_msg, reply_markup=keyboard)
    except ValueError:
        result_msg = "❌ User ID raqam bo'lishi kerak"
        await message.answer(result_msg, reply_markup=keyboard)
    
    await state.clear()


@admin_router.callback_query(F.data == "admin_help")
async def admin_help_callback(query: CallbackQuery):
    """Show admin help"""
    if not db.is_admin(query.from_user.id):
        await query.answer("❌ Siz admin emassiz!", show_alert=True)
        return
    
    help_text = r"""
🔐 ADMIN BUYRUQLARI

/admin - Admin menyu ko'rsatish
/stats - Bugungi statistika
/reports - Bugungi hisobotlar
/interns_list - Barcha talabalar
/search_report [Ism] - Talabani qidirish
/delete_report [Ism] [DD.MM.YYYY] - Hisobotni o'chirish
/admin_add [User_ID] - Admin qo'shish
/logs - Faoliyat jurnali
/export_excel - Excel ga eksport

MISOL:
/search_report Humoyun Jo'rayev
/delete_report Humoyun\ Jo\'rayev 17.04.2026
/admin_add 123456789
"""
    
    await query.message.edit_text(help_text, reply_markup=get_back_keyboard())
    await query.answer()


@admin_router.message(F.text == "/interns_list")
@admin_router.message(F.text == "👥 Talabalar")
@admin_router.message(F.text == "Internlar ro'yxati")
@admin_router.message(F.text == "internlar ro'yxati")
async def admin_interns_list(message: Message):
    """Show all interns list"""
    if not db.is_admin(message.from_user.id):
        await message.answer("❌ Siz admin emassiz!")
        return
    
    interns_text = "👥 BARCHA TALABALAR\n\n"
    for i, intern in enumerate(INTERNS, 1):
        interns_text += f"{i}. {intern}\n"
    
    await message.answer(interns_text)
    db.add_log(message.from_user.id, "interns_list_viewed")


@admin_router.message(F.text.startswith("/search_report"))
async def admin_search_report(message: Message, state: FSMContext):
    """Search for specific report (legacy command handler)"""
    if not db.is_admin(message.from_user.id):
        await message.answer("❌ Siz admin emassiz!")
        return
    
    parts = message.text.split(" ", 1)
    if len(parts) < 2:
        await state.set_state(AdminStates.searching_intern)
        await message.answer("🔍 Talabaning ismi ni yozing:")
        return
    
    intern_name = parts[1]
    reports = db.get_reports_by_intern(intern_name, days=30)
    
    if not reports:
        await message.answer(f"❌ {intern_name} uchun hisobotlar topilmadi")
        return
    
    search_text = f"🔍 {intern_name} - Oxirgi 30 kun hisobotlari:\n\n"
    
    for report in reports:
        status = "✅" if report['status'] == 'Keldi' else "❌"
        search_text += f"{status} {report['report_date']}\n"
        
        if report['status'] == 'Keldi':
            search_text += f"   {report['arrival_time']} - {report['departure_time']} | {report['lesson_count']} dars\n"
        else:
            search_text += f"   {report['absence_reason']}\n"
        search_text += "\n"
    
    await message.answer(search_text)
    db.add_log(message.from_user.id, "search_report", f"Intern: {intern_name}")


@admin_router.message(F.text.startswith("/delete_report"))
async def admin_delete_report(message: Message, state: FSMContext):
    """Delete a report (legacy command handler)"""
    if not db.is_admin(message.from_user.id):
        await message.answer("❌ Siz admin emassiz!")
        return
    
    parts = message.text.split(" ", 2)
    if len(parts) < 3:
        await state.set_state(AdminStates.deleting_report)
        await message.answer("🗑️ Talabaning ismi va sanani yozing:\nMisol: Humoyun 17.04.2026")
        return
    
    intern_name = parts[1]
    date_str = parts[2]
    
    try:
        report_date = date.strptime(date_str, "%d.%m.%Y")
        if db.delete_report(intern_name, report_date):
            await message.answer(f"✅ Hisobot o'chirildi: {intern_name} ({date_str})")
            db.add_log(message.from_user.id, "report_deleted", f"{intern_name} on {date_str}")
        else:
            await message.answer("❌ Hisobot o'chirilmadi")
    except ValueError:
        await message.answer("❌ Sana formati xato (DD.MM.YYYY ishlatilsin)")


@admin_router.message(F.text.startswith("/admin_add"))
async def admin_add(message: Message, state: FSMContext):
    """Add new admin user (legacy command handler)"""
    if not db.is_admin(message.from_user.id):
        await message.answer("❌ Siz admin emassiz!")
        return
    
    parts = message.text.split(" ", 1)
    if len(parts) < 2:
        await state.set_state(AdminStates.adding_admin)
        await message.answer("👨‍💼 Yangi admin User ID sini yozing:")
        return
    
    try:
        new_admin_id = int(parts[1])
        if db.add_admin(new_admin_id):
            await message.answer(f"✅ Admin qo'shildi: {new_admin_id}")
            db.add_log(message.from_user.id, "admin_added", f"Admin ID: {new_admin_id}")
        else:
            await message.answer("❌ Admin qo'shilmadi (Balki allaqachon admin)")
    except ValueError:
        await message.answer("❌ User ID raqam bo'lishi kerak")


@admin_router.message(F.text == "/logs")
async def admin_logs(message: Message):
    """Show activity logs (legacy command handler)"""
    if not db.is_admin(message.from_user.id):
        await message.answer("❌ Siz admin emassiz!")
        return
    
    logs = db.get_logs(limit=20)
    
    logs_text = "📜 FAOLIYAT JURNALI (Oxirgi 20)\n\n"
    
    for log in logs:
        logs_text += f"👤 {log['user_id']}\n"
        logs_text += f"   ▪ {log['action']}\n"
        if log['details']:
            logs_text += f"   ▪ {log['details']}\n"
        logs_text += f"   ⏰ {log['created_at']}\n\n"
    
    await message.answer(logs_text)


@admin_router.message(F.text == "/export_excel")
async def admin_export_excel(message: Message):
    """Export data to Excel (legacy command handler)"""
    if not db.is_admin(message.from_user.id):
        await message.answer("❌ Siz admin emassiz!")
        return
    
    try:
        import pandas as pd
        from database import DATABASE_FILE
        
        lessons = db.get_all_lessons(days=90)
        
        if not lessons:
            await message.answer("❌ Export uchun darslar topilmadi")
            return
        
        # Prepare data for Excel
        data = []
        for lesson in lessons:
            data.append({
                'Sana': lesson['lesson_date'],
                'ISM': lesson['intern_name'],
                'Dars #': lesson['lesson_number'],
                'Ustoz': lesson['teacher_name'],
                'Xona': lesson['room'],
                'Vaqt': lesson['lesson_time']
            })
        
        df = pd.DataFrame(data)
        
        # Save to Excel
        excel_path = DATABASE_FILE.parent / f"lessons_export_{date.today().strftime('%d_%m_%Y')}.xlsx"
        df.to_excel(excel_path, index=False, sheet_name='Darslar')
        
        await message.answer(f"✅ Eksport bajarildi: {excel_path.name}")
        db.add_log(message.from_user.id, "excel_exported")
    except Exception as e:
        await message.answer(f"❌ Eksport xatosi: {str(e)}")


@admin_router.message(F.text == "/admin_help")
async def admin_help(message: Message):
    """Show admin help"""
    if not db.is_admin(message.from_user.id):
        await message.answer("❌ Siz admin emassiz!")
        return
    
    help_text = r"""
🔐 ADMIN BUYRUQLARI

/admin - Admin menyu ko'rsatish
/stats - Bugungi statistika
/reports - Bugungi hisobotlar
/interns_list - Barcha talabalar
/search_report [Ism] - Talabani qidirish
/delete_report [Ism] [DD.MM.YYYY] - Hisobotni o'chirish
/admin_add [User_ID] - Admin qo'shish
/logs - Faoliyat jurnali
/export_excel - Excel ga eksport

MISOL:
/search_report Humoyun Jo'rayev
/delete_report Humoyun\ Jo\'rayev 17.04.2026
/admin_add 123456789
"""
    
    await message.answer(help_text)


@admin_router.callback_query(F.data == "admin_back")
async def admin_back_callback(query: CallbackQuery):
    """Go back to admin menu"""
    if not db.is_admin(query.from_user.id):
        await query.answer("❌ Siz admin emassiz!", show_alert=True)
        return
    
    # Create inline keyboard with buttons
    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="📊 Statistika", callback_data="admin_stats"),
            InlineKeyboardButton(text="📋 Hisobotlar", callback_data="admin_reports")
        ],
        [
            InlineKeyboardButton(text="📚 Dars ro'yxati", callback_data="admin_lessons"),
            InlineKeyboardButton(text="👥 Talabalar", callback_data="admin_interns")
        ],
        [
            InlineKeyboardButton(text="⏱️ Ish vaqti", callback_data="admin_work_stats"),
            InlineKeyboardButton(text="📥 Excel export", callback_data="admin_excel")
        ],
        [
            InlineKeyboardButton(text="📜 Jurnali", callback_data="admin_logs"),
            InlineKeyboardButton(text="🔍 Qidirish", callback_data="admin_search")
        ],
        [
            InlineKeyboardButton(text="🗑️ O'chirish", callback_data="admin_delete"),
            InlineKeyboardButton(text="👨‍💼 Admin qo'shish", callback_data="admin_add")
        ],
        [
            InlineKeyboardButton(text="ℹ️ Yordam", callback_data="admin_help")
        ]
    ])
    
    menu_text = """
🔐 ADMIN PANEL

Siz quyidagi amalarni bajarishingiz mumkin:
1️⃣ 📊 Statistika - Bugungi statistikani ko'rish
2️⃣ 📋 Hisobotlar - Bugungi hisobotlarni ko'rish
3️⃣ 📚 Dars ro'yxati - Interns kiritgan darslarni ko'rish
4️⃣ 👥 Talabalar - Barcha talabalari ko'rish
5️⃣ ⏱️ Ish vaqti - Har bir intern nechi soat marsda bo'gani
6️⃣ 📥 Excel export - Darslarni Excel ga chiqarish
7️⃣ 📜 Jurnali - Faoliyat jurnalini ko'rish
8️⃣ 🔍 Qidirish - Talabani qidirish
9️⃣ 🗑️ O'chirish - Hisobotni o'chirish
"""
    
    await query.message.edit_text(menu_text, reply_markup=keyboard)
    await query.answer()
