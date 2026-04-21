"""
Excel handler for storing and managing intern reports with beautiful formatting
"""
from pathlib import Path
from datetime import date
from typing import Dict, List, Optional
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config import EXCEL_FILE
from interns import INTERNS


class ExcelHandler:
    """Handle Excel operations for intern reports"""

    def __init__(self, excel_file: Path = EXCEL_FILE):
        self.excel_file = excel_file
        self.ensure_file_exists()

    def ensure_file_exists(self):
        """Create Excel file if it doesn't exist"""
        if not self.excel_file.exists():
            self.create_new_file()

    def create_new_file(self):
        """Create a new Excel file with headers and intern names"""
        data = {
            'Ism Familiya': INTERNS,
        }
        df = pd.DataFrame(data)
        
        # Create excel file
        self.excel_file.parent.mkdir(parents=True, exist_ok=True)
        df.to_excel(self.excel_file, sheet_name='Hisobot', index=False)
        
        self.setup_styles()

    def setup_styles(self):
        """Setup Excel styling with professional formatting"""
        wb = load_workbook(self.excel_file)
        ws = wb.active
        
        # Define professional colors
        header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')  # Dark blue
        header_font = Font(bold=True, color='FFFFFF', size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply header styling
        for col_idx, cell in enumerate(ws[1], start=1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        # Apply alternating row colors and borders
        light_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')  # Light gray
        white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')  # White
        body_font = Font(color='000000', size=11)
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            fill = light_fill if row_idx % 2 == 0 else white_fill
            for cell in row:
                cell.fill = fill
                cell.font = body_font
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.border = border
        
        # Set column widths
        ws.column_dimensions['A'].width = 30
        
        # Freeze first column and header
        ws.freeze_panes = 'B2'
        
        # Set default row height
        ws.row_dimensions[1].height = 25
        for row_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 30
        
        wb.save(self.excel_file)
        wb.close()

    def add_record(self, data: Dict):
        """
        Add or update a record in Excel
        
        data structure:
        {
            'intern_name': str,
            'date': date,
            'arrival_time': str,
            'departure_time': str,
            'lessons': list,
            'status': 'Keldi' or 'Kelmadi',
            'absence_reason': str (optional)
        }
        """
        df = pd.read_excel(self.excel_file, sheet_name='Hisobot')
        
        # Find or create date column
        date_str = data['date'].strftime('%d.%m.%Y')
        
        if date_str not in df.columns:
            df[date_str] = ''
        
        # Find intern row
        intern_row = df[df['Ism Familiya'] == data['intern_name']]
        
        if intern_row.empty:
            # Add new intern (shouldn't happen, but just in case)
            new_row = {'Ism Familiya': data['intern_name'], date_str: ''}
            df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
        
        # Prepare record value
        if data['status'] == 'Keldi':
            lesson_count = len(data['lessons'])
            teachers = ', '.join([lesson['teacher'] for lesson in data['lessons']])
            record_value = f"[Keldi] {lesson_count} dars\n{teachers}"
        else:
            reason = data.get('absence_reason', 'Sabab ko\'rsatilmadi')
            record_value = f"[Kelmadi] {reason}"
        
        # Update record
        df.loc[df['Ism Familiya'] == data['intern_name'], date_str] = record_value
        
        # Save to Excel
        df.to_excel(self.excel_file, sheet_name='Hisobot', index=False)
        
        # Apply styles and coloring
        self.apply_formatting(data, date_str)

    def apply_formatting(self, data: Dict, date_str: str):
        """Apply professional formatting to cells"""
        wb = load_workbook(self.excel_file)
        ws = wb.active
        
        # Define borders
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Find intern row
        intern_row = None
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if row[0].value == data['intern_name']:
                intern_row = row_idx
                break
        
        if not intern_row:
            wb.close()
            return
        
        # Find column by date
        date_col = None
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value == date_str:
                date_col = col_idx
                break
        
        if not date_col:
            wb.close()
            return
        
        # Get cell and apply formatting
        cell = ws.cell(row=intern_row, column=date_col)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
        
        if data['status'] == 'Kelmadi':
            # Red for absent - missing
            red_fill = PatternFill(start_color='F4CCCC', end_color='F4CCCC', fill_type='solid')  # Light red
            red_font = Font(color='900000', bold=True, size=11)  # Dark red text
            cell.fill = red_fill
            cell.font = red_font
        else:
            # Green for present - came
            green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Light green
            green_font = Font(color='006100', bold=True, size=11)  # Dark green text
            cell.fill = green_fill
            cell.font = green_font
        
        # Set row height
        ws.row_dimensions[intern_row].height = 35
        
        wb.save(self.excel_file)
        wb.close()

    def check_today_attendance(self, today: date) -> List[str]:
        """Get list of interns who haven't submitted report today"""
        df = pd.read_excel(self.excel_file, sheet_name='Hisobot')
        
        date_str = today.strftime('%d.%m.%Y')
        
        if date_str not in df.columns:
            # No reports submitted today yet
            return INTERNS.copy()
        
        # Find interns who submitted
        submitted = df[df[date_str].notna() & (df[date_str] != '')]
        submitted_names = submitted['Ism Familiya'].tolist()
        
        # Return interns who didn't submit
        not_submitted = [name for name in INTERNS if name not in submitted_names]
        return not_submitted

    def get_attendance_summary(self, target_date: date = None) -> Dict:
        """Get attendance summary for a specific date"""
        if target_date is None:
            target_date = date.today()
        
        df = pd.read_excel(self.excel_file, sheet_name='Hisobot')
        date_str = target_date.strftime('%d.%m.%Y')
        
        if date_str not in df.columns:
            return {
                'date': target_date,
                'present': 0,
                'absent': 0,
                'total': len(INTERNS),
                'details': {}
            }
        
        present = 0
        absent = 0
        details = {}
        
        for _, row in df.iterrows():
            name = row['Ism Familiya']
            status = row[date_str]
            
            if pd.isna(status) or status == '':
                details[name] = 'No report'
            elif '[Keldi]' in str(status):
                present += 1
                details[name] = 'Present'
            elif '[Kelmadi]' in str(status):
                absent += 1
                details[name] = 'Absent'
        
        return {
            'date': target_date,
            'present': present,
            'absent': absent,
            'total': len(INTERNS),
            'not_reported': len(INTERNS) - present - absent,
            'details': details
        }
